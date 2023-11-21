import googlemaps
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
from dotenv import find_dotenv, load_dotenv

# Load environment variables
load_dotenv(find_dotenv())

# Variable Declaration
api_key = os.environ["GOOGLE_MAPS_API_KEY"]
file_path = os.environ["FILE"]

# Initialize Google Maps Client
gMaps = googlemaps.Client(key=api_key)

# Load the workbook and get the active sheet
wb = load_workbook(filename=file_path)	
sheet = wb.active

# Iterate over the rows and update longitude and latitude
for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column, values_only=False):
    if(row[7].value == 0):
        address = row[3].value + ", " + row[4].value + ", Jamaica"
        if address:
            geocode_result = gMaps.geocode(address)
            if geocode_result:
                latitude = geocode_result[0]["geometry"]["location"]["lat"]
                longitude = geocode_result[0]["geometry"]["location"]["lng"]
                row[5].value = longitude
                row[6].value = latitude

# Save the workbook
wb.save(filename=file_path)